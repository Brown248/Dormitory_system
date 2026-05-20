from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
import os
from dotenv import load_dotenv
import re
import io
from datetime import datetime, date, time, timedelta
import models, schemas, database
from auth import create_token, LoginRequest
from database import engine, get_db
from linebot.v3 import WebhookParser
from linebot.v3.messaging import Configuration, AsyncApiClient, AsyncMessagingApi, ReplyMessageRequest, TextMessage
from linebot.v3.exceptions import InvalidSignatureError
from linebot.v3.webhooks import MessageEvent, TextMessageContent

load_dotenv()

models.Base.metadata.create_all(bind=engine)

# LINE Config
channel_secret = os.getenv('LINE_CHANNEL_SECRET')
channel_access_token = os.getenv('LINE_CHANNEL_ACCESS_TOKEN')
parser = WebhookParser(channel_secret)

_line_bot_api = None

def get_line_bot_api() -> AsyncMessagingApi:
    global _line_bot_api
    if _line_bot_api is None:
        configuration = Configuration(access_token=channel_access_token)
        async_api_client = AsyncApiClient(configuration)
        _line_bot_api = AsyncMessagingApi(async_api_client)
    return _line_bot_api


from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="Sovereign Accounting API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# Auth Endpoint
# ==========================================
@app.post("/auth/login")
def login(req: LoginRequest):
    admin_password = os.getenv('ADMIN_PASSWORD', 'horpak2026')
    if req.password != admin_password:
        raise HTTPException(status_code=401, detail="รหัสผ่านไม่ถูกต้อง")
    token = create_token({"role": "admin"})
    return {"access_token": token, "token_type": "bearer"}

@app.on_event("startup")
def seed_business_units():
    db = next(get_db())
    try:
        if db.query(models.BusinessUnit).count() == 0:
            units = [
                {"name": "หอพัก", "type": models.UnitType.DORMITORY},
                {"name": "อู่ซ่อมรถ", "type": models.UnitType.GARAGE},
                {"name": "บ้านเช่า หลังที่ 1", "type": models.UnitType.HOUSE},
                {"name": "บ้านเช่า หลังที่ 2", "type": models.UnitType.HOUSE},
                {"name": "บ้านเช่า หลังที่ 3", "type": models.UnitType.HOUSE},
            ]
            for unit in units:
                db.add(models.BusinessUnit(**unit))
            db.commit()
            print("Successfully seeded 5 default Business Units.")
    except Exception as e:
        db.rollback()
        print(f"Error seeding business units: {e}")
    finally:
        db.close()
    
    # Seed DormRooms and Houses
    seed_rooms_and_houses()

def seed_rooms_and_houses():
    db = next(get_db())
    try:
        # Seed DormRooms
        if db.query(models.DormRoom).count() == 0:
            rooms = []
            
            # 26_20
            rooms_26_20 = [
                {"number": "401", "rate": 2800, "floor": 4}, {"number": "402", "rate": 2500, "floor": 4},
                {"number": "403", "rate": 2500, "floor": 4}, {"number": "404", "rate": 2200, "floor": 4},
                {"number": "405", "rate": 2500, "floor": 4}, {"number": "406", "rate": 2500, "floor": 4},
                {"number": "301", "rate": 2800, "floor": 3}, {"number": "302", "rate": 2500, "floor": 3},
                {"number": "303", "rate": 2500, "floor": 3}, {"number": "304", "rate": 3000, "floor": 3},
                {"number": "305", "rate": 2800, "floor": 3}, {"number": "306", "rate": 2800, "floor": 3},
                {"number": "307", "rate": 3000, "floor": 3},
                {"number": "201", "rate": 3000, "floor": 2}, {"number": "202", "rate": 2800, "floor": 2},
                {"number": "203", "rate": 2800, "floor": 2}, {"number": "204", "rate": 3000, "floor": 2},
                {"number": "205", "rate": 2800, "floor": 2}, {"number": "206", "rate": 2500, "floor": 2},
                {"number": "207", "rate": 2700, "floor": 2},
                {"number": "101", "rate": 2500, "floor": 1}, {"number": "102", "rate": 2500, "floor": 1},
                {"number": "103", "rate": 2500, "floor": 1}, {"number": "104", "rate": 2500, "floor": 1},
            ]
            for r in rooms_26_20:
                rooms.append(models.DormRoom(
                    dorm_key="26_20",
                    number=r["number"],
                    floor=r["floor"],
                    rate=float(r["rate"]),
                    tenant="",
                    payment_status="pending"
                ))

            # 26_577
            # Floor 3: 301 to 310
            for i in range(1, 11):
                rooms.append(models.DormRoom(
                    dorm_key="26_577",
                    number=f"3{str(i).zfill(2)}",
                    floor=3,
                    rate=2500.0,
                    tenant="",
                    payment_status="pending"
                ))
            # Floor 2: 201 to 210
            for i in range(1, 11):
                rooms.append(models.DormRoom(
                    dorm_key="26_577",
                    number=f"2{str(i).zfill(2)}",
                    floor=2,
                    rate=2500.0,
                    tenant="",
                    payment_status="pending"
                ))
            # Floor 1
            rooms_26_577_f1 = [
                {"number": "101", "rate": 2500, "floor": 1}, {"number": "102", "rate": 2500, "floor": 1},
                {"number": "103", "rate": 2500, "floor": 1}, {"number": "104", "rate": 2000, "floor": 1},
                {"number": "105", "rate": 2500, "floor": 1}, {"number": "106", "rate": 2500, "floor": 1},
                {"number": "107", "rate": 2500, "floor": 1}, {"number": "108", "rate": 2500, "floor": 1},
                {"number": "109", "rate": 2500, "floor": 1}, {"number": "110", "rate": 2500, "floor": 1},
            ]
            for r in rooms_26_577_f1:
                rooms.append(models.DormRoom(
                    dorm_key="26_577",
                    number=r["number"],
                    floor=r["floor"],
                    rate=float(r["rate"]),
                    tenant="",
                    payment_status="pending"
                ))

            # 73_17
            # Floor 2: B1 to B9
            for i in range(1, 10):
                rooms.append(models.DormRoom(
                    dorm_key="73_17",
                    number=f"B{i}",
                    floor=2,
                    rate=3500.0,
                    tenant="",
                    payment_status="pending"
                ))
            # Floor 1: A1 to A8
            for i in range(1, 9):
                rooms.append(models.DormRoom(
                    dorm_key="73_17",
                    number=f"A{i}",
                    floor=1,
                    rate=3500.0,
                    tenant="",
                    payment_status="pending"
                ))

            for r in rooms:
                db.add(r)
            db.commit()
            print(f"Successfully seeded {len(rooms)} Dormitory Rooms.")

        # Seed RentalHouses
        if db.query(models.RentalHouse).count() == 0:
            houses = [
                models.RentalHouse(id="h1", name="บ้านเช่า หลังที่ 1", tenant_name="", monthly_rent=5000.0, payment_status="unpaid"),
                models.RentalHouse(id="h2", name="บ้านเช่า หลังที่ 2", tenant_name="", monthly_rent=4500.0, payment_status="unpaid"),
                models.RentalHouse(id="h3", name="บ้านเช่า หลังที่ 3", tenant_name="", monthly_rent=6000.0, payment_status="unpaid"),
            ]
            for h in houses:
                db.add(h)
            db.commit()
            print("Successfully seeded 3 default Rental Houses.")

    except Exception as e:
        db.rollback()
        print(f"Error seeding rooms and houses: {e}")
    finally:
        db.close()

@app.get("/")
def read_root():
    return {"message": "Welcome to Sovereign Accounting API"}

# Business Units
@app.post("/units/", response_model=schemas.BusinessUnit)
def create_unit(unit: schemas.BusinessUnitCreate, db: Session = Depends(get_db)):
    db_unit = models.BusinessUnit(**unit.dict())
    db.add(db_unit)
    db.commit()
    db.refresh(db_unit)
    return db_unit

@app.get("/units/", response_model=List[schemas.BusinessUnit])
def read_units(db: Session = Depends(get_db)):
    return db.query(models.BusinessUnit).all()

# Customers
@app.post("/customers/", response_model=schemas.Customer)
def create_customer(customer: schemas.CustomerCreate, db: Session = Depends(get_db)):
    db_customer = models.Customer(**customer.dict())
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer

@app.get("/customers/", response_model=List[schemas.Customer])
def read_customers(db: Session = Depends(get_db)):
    return db.query(models.Customer).all()

# Invoices
@app.post("/invoices/", response_model=schemas.Invoice)
def create_invoice(invoice: schemas.InvoiceCreate, db: Session = Depends(get_db)):
    db_invoice = models.Invoice(**invoice.dict())
    db.add(db_invoice)
    db.commit()
    db.refresh(db_invoice)
    # TODO: Trigger LINE Notification (Push Message)
    return db_invoice

@app.get("/invoices/", response_model=List[schemas.Invoice])
def read_invoices(db: Session = Depends(get_db)):
    return db.query(models.Invoice).all()

@app.patch("/invoices/{invoice_id}/status", response_model=schemas.Invoice)
def update_invoice_status(invoice_id: int, status: str, db: Session = Depends(get_db)):
    db_invoice = db.query(models.Invoice).filter(models.Invoice.id == invoice_id).first()
    if not db_invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    try:
        # Convert string to Enum
        db_invoice.status = models.InvoiceStatus(status.lower())
        db.commit()
        db.refresh(db_invoice)
        
        # When an invoice is paid, we can automatically record it as an INCOME transaction!
        if status.lower() == "paid":
            # Check if this invoice already has a matching transaction to prevent duplicate ledger records
            existing_tx = db.query(models.Transaction).filter(
                models.Transaction.description.like(f"%บิล ID: {db_invoice.id}%")
            ).first()
            if not existing_tx:
                db_tx = models.Transaction(
                    type=models.TransactionType.INCOME,
                    amount=db_invoice.amount,
                    description=f"รับชำระบิล: {db_invoice.title} (บิล ID: {db_invoice.id})",
                    unit_id=db_invoice.unit_id,
                    customer_id=db_invoice.customer_id
                )
                db.add(db_tx)
                db.commit()
        
        return db_invoice
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid invoice status")

# Transactions
@app.post("/transactions/", response_model=schemas.Transaction)
def create_transaction(transaction: schemas.TransactionCreate, db: Session = Depends(get_db)):
    db_transaction = models.Transaction(**transaction.dict())
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    return db_transaction

@app.get("/transactions/", response_model=List[schemas.Transaction])
def read_transactions(db: Session = Depends(get_db)):
    return db.query(models.Transaction).order_by(models.Transaction.created_at.desc()).all()

from sqlalchemy import func
from typing import List, Optional

@app.get("/transactions/summary", response_model=schemas.DashboardSummary)
def get_transaction_summary(db: Session = Depends(get_db)):
    # SQL-level aggregation for global totals
    totals = db.query(
        models.Transaction.type,
        func.sum(models.Transaction.amount)
    ).group_by(models.Transaction.type).all()
    
    total_income = next((float(t[1]) for t in totals if t[0] == models.TransactionType.INCOME), 0.0)
    total_expense = next((float(t[1]) for t in totals if t[0] == models.TransactionType.EXPENSE), 0.0)
    balance = total_income - total_expense
    
    # SQL-level aggregation for unit breakdowns
    unit_totals = db.query(
        models.Transaction.unit_id,
        models.Transaction.type,
        func.sum(models.Transaction.amount)
    ).group_by(models.Transaction.unit_id, models.Transaction.type).all()
    
    units = db.query(models.BusinessUnit).all()
    unit_summaries = []
    
    for u in units:
        u_income = sum(float(t[2]) for t in unit_totals if t[0] == u.id and t[1] == models.TransactionType.INCOME)
        u_expense = sum(float(t[2]) for t in unit_totals if t[0] == u.id and t[1] == models.TransactionType.EXPENSE)
        unit_summaries.append(
            schemas.BusinessUnitSummary(
                id=u.id,
                name=u.name,
                type=u.type,
                total_income=u_income,
                total_expense=u_expense,
                balance=u_income - u_expense
            )
        )
        
    return schemas.DashboardSummary(
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        units=unit_summaries
    )

# ==========================================
# Export Excel Endpoint
# ==========================================
@app.get("/transactions/export/excel")
def export_transactions_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    transactions = db.query(models.Transaction).order_by(models.Transaction.created_at.desc()).all()
    units = db.query(models.BusinessUnit).all()
    unit_map = {u.id: u.name for u in units}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายการบัญชี"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["ลำดับ", "วันที่", "ประเภท", "จำนวนเงิน (บาท)", "รายละเอียด", "ธุรกิจ"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    income_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    expense_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    for idx, tx in enumerate(transactions, 1):
        row = idx + 1
        tx_type = "รายรับ" if tx.type == models.TransactionType.INCOME else "รายจ่าย"
        unit_name = unit_map.get(tx.unit_id, "ทั่วไป")
        date_str = tx.created_at.strftime("%d/%m/%Y %H:%M") if tx.created_at else ""
        fill = income_fill if tx.type == models.TransactionType.INCOME else expense_fill
        
        data = [idx, date_str, tx_type, float(tx.amount), tx.description, unit_name]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.fill = fill
            if col == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
    
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)
    
    total_income = sum(float(tx.amount) for tx in transactions if tx.type == models.TransactionType.INCOME)
    total_expense = sum(float(tx.amount) for tx in transactions if tx.type == models.TransactionType.EXPENSE)
    last_row = len(transactions) + 2
    ws.cell(row=last_row, column=3, value="รวมรายรับ:").font = Font(bold=True)
    ws.cell(row=last_row, column=4, value=total_income).font = Font(bold=True, color="228B22")
    ws.cell(row=last_row, column=4).number_format = '#,##0.00'
    ws.cell(row=last_row+1, column=3, value="รวมรายจ่าย:").font = Font(bold=True)
    ws.cell(row=last_row+1, column=4, value=total_expense).font = Font(bold=True, color="FF0000")
    ws.cell(row=last_row+1, column=4).number_format = '#,##0.00'
    ws.cell(row=last_row+2, column=3, value="คงเหลือสุทธิ:").font = Font(bold=True)
    ws.cell(row=last_row+2, column=4, value=total_income - total_expense).font = Font(bold=True, color="1F4E79")
    ws.cell(row=last_row+2, column=4).number_format = '#,##0.00'
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"sovereign_transactions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==========================================
# Monthly Summary Endpoint (for Charts)
# ==========================================
@app.get("/transactions/monthly-summary")
def get_monthly_summary(db: Session = Depends(get_db)):
    six_months_ago = datetime.now() - timedelta(days=180)
    
    transactions = db.query(models.Transaction).filter(
        models.Transaction.created_at >= six_months_ago
    ).all()
    
    monthly_data = {}
    for tx in transactions:
        if tx.created_at:
            key = tx.created_at.strftime("%Y-%m")
            if key not in monthly_data:
                monthly_data[key] = {"month": key, "income": 0.0, "expense": 0.0}
            if tx.type == models.TransactionType.INCOME:
                monthly_data[key]["income"] += float(tx.amount)
            else:
                monthly_data[key]["expense"] += float(tx.amount)
    
    result = sorted(monthly_data.values(), key=lambda x: x["month"])
    return result[-6:] if len(result) > 6 else result

def match_business_unit(description: str, db: Session) -> Optional[models.BusinessUnit]:
    units = db.query(models.BusinessUnit).all()
    if not units:
        return None
    desc_lower = description.lower()
    
    # 1. Garage keywords
    if any(k in desc_lower for k in ["อู่", "ซ่อม", "garage", "รถ", "ตู้"]):
        for u in units:
            if u.type == models.UnitType.GARAGE:
                return u
                
    # 2. House keywords
    if any(k in desc_lower for k in ["บ้าน", "house"]):
        if any(k in desc_lower for k in ["1", "หลังที่ 1", "หลัง 1"]):
            for u in units:
                if "หลังที่ 1" in u.name:
                    return u
        elif any(k in desc_lower for k in ["2", "หลังที่ 2", "หลัง 2"]):
            for u in units:
                if "หลังที่ 2" in u.name:
                    return u
        elif any(k in desc_lower for k in ["3", "หลังที่ 3", "หลัง 3"]):
            for u in units:
                if "หลังที่ 3" in u.name:
                    return u
        # Fallback to first house if "บ้าน" matched but no specific number
        for u in units:
            if u.type == models.UnitType.HOUSE:
                return u
                
    # 3. Dormitory keywords
    if any(k in desc_lower for k in ["หอ", "dorm", "ห้อง", "b20", "b577", "a20"]):
        for u in units:
            if u.type == models.UnitType.DORMITORY:
                return u
                
    return None # Return None instead of units[0] to allow caller to handle "General" status

@app.post("/webhook")
async def line_webhook(request: Request, db: Session = Depends(get_db)):
    signature = request.headers.get('X-Line-Signature')
    body = await request.body()
    body_str = body.decode('utf-8')

    try:
        events = parser.parse(body_str, signature)
    except InvalidSignatureError:
        raise HTTPException(status_code=400, detail="Invalid signature")

    for event in events:
        if not isinstance(event, MessageEvent) or not isinstance(event.message, TextMessageContent):
            continue
        
        user_id = event.source.user_id
        text = event.message.text.strip()

        # Command A: เช็คยอด
        if text == "เช็คยอด":
            customer = db.query(models.Customer).filter(models.Customer.line_user_id == user_id).first()
            if not customer:
                reply_text = "ขออภัยครับ ไม่พบข้อมูลของคุณในระบบ กรุณาแจ้งแอดมินเพื่อผูกบัญชี LINE ครับ"
            else:
                unpaid_invoices = [i for i in customer.invoices if i.status == models.InvoiceStatus.UNPAID]
                if not unpaid_invoices:
                    reply_text = f"สวัสดีครับคุณ {customer.name}\nคุณไม่มียอดค้างชำระครับ"
                else:
                    total = sum(i.amount for i in unpaid_invoices)
                    details = "\n".join([f"- {i.title}: {i.amount:,.2f} บาท" for i in unpaid_invoices])
                    reply_text = f"สวัสดีครับคุณ {customer.name}\nยอดค้างชำระทั้งหมด: {total:,.2f} บาท\n\nรายละเอียด:\n{details}"
            
            await get_line_bot_api().reply_message(
                ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=reply_text)]
                )
            )
            continue

        # Command B: สรุปวันนี้
        if text == "สรุปวันนี้":
            today_start = datetime.combine(date.today(), time.min)
            today_end = datetime.combine(date.today(), time.max)
            
            today_tx = db.query(models.Transaction).filter(
                models.Transaction.created_at >= today_start,
                models.Transaction.created_at <= today_end
            ).all()
            
            total_inc = sum(t.amount for t in today_tx if t.type == models.TransactionType.INCOME)
            total_exp = sum(t.amount for t in today_tx if t.type == models.TransactionType.EXPENSE)
            net_bal = total_inc - total_exp
            
            reply_text = (
                f"📅 สรุปยอดบัญชีวันนี้\n"
                f"-----------------------\n"
                f"📈 รายรับรวม: +{total_inc:,.2f} บาท\n"
                f"📉 รายจ่ายรวม: -{total_exp:,.2f} บาท\n"
                f"-----------------------\n"
                f"💰 คงเหลือสุทธิ: {net_bal:,.2f} บาท\n\n"
                f"(บันทึกผ่าน LINE ทั้งหมด {len(today_tx)} รายการ)"
            )
            
            await get_line_bot_api().reply_message(
                ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=reply_text)]
                )
            )
            continue

        # Command C: ยอดเงิน
        if text == "ยอดเงิน":
            units = db.query(models.BusinessUnit).all()
            transactions = db.query(models.Transaction).all()
            
            total_inc_all = sum(t.amount for t in transactions if t.type == models.TransactionType.INCOME)
            total_exp_all = sum(t.amount for t in transactions if t.type == models.TransactionType.EXPENSE)
            net_all = total_inc_all - total_exp_all
            
            lines = []
            for u in units:
                u_tx = [t for t in transactions if t.unit_id == u.id]
                u_inc = sum(t.amount for t in u_tx if t.type == models.TransactionType.INCOME)
                u_exp = sum(t.amount for t in u_tx if t.type == models.TransactionType.EXPENSE)
                u_bal = u_inc - u_exp
                lines.append(f"🏢 {u.name}: {u_bal:,.2f} บาท")
                
            details = "\n".join(lines)
            reply_text = (
                f"💰 ยอดเงินคงเหลือแต่ละธุรกิจ\n"
                f"-----------------------\n"
                f"{details}\n"
                f"-----------------------\n"
                f"💵 ยอดรวมทั้งหมด: {net_all:,.2f} บาท"
            )
            
            await get_line_bot_api().reply_message(
                ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=reply_text)]
                )
            )
            continue

        # Command E: ทวงบิล
        if text == "ทวงบิล":
            unpaid_rooms = db.query(models.DormRoom).filter(
                models.DormRoom.payment_status != "paid",
                models.DormRoom.tenant != "",
                models.DormRoom.tenant != None
            ).all()
            unpaid_houses = db.query(models.RentalHouse).filter(
                models.RentalHouse.payment_status != "paid",
                models.RentalHouse.tenant_name != "",
                models.RentalHouse.tenant_name != None
            ).all()
            unpaid_jobs = db.query(models.GarageJob).filter(
                models.GarageJob.payment_status != "paid"
            ).all()
            
            lines = ["\U0001f514 สรุปบิลค้างชำระ"]
            lines.append("=======================")
            
            if unpaid_rooms:
                lines.append(f"\n\U0001f3e2 หอพัก ({len(unpaid_rooms)} ห้อง):")
                for r in unpaid_rooms[:15]:
                    total = r.rate + r.water_cost + r.electric_cost + r.cleaning_fee + r.other_fee + r.fine_cost
                    lines.append(f"  ห้อง {r.number} ({r.dorm_key}) - {r.tenant}: {total:,.0f}\u0e3f")
            
            if unpaid_houses:
                lines.append(f"\n\U0001f3e0 บ้านเช่า ({len(unpaid_houses)} หลัง):")
                for h in unpaid_houses:
                    total = h.monthly_rent + h.water_bill + h.electric_bill
                    lines.append(f"  {h.name} - {h.tenant_name}: {total:,.0f}\u0e3f")
            
            if unpaid_jobs:
                lines.append(f"\n\U0001f527 อู่ซ่อมรถ ({len(unpaid_jobs)} งาน):")
                for j in unpaid_jobs[:10]:
                    lines.append(f"  {j.license_plate} ({j.customer_name}): {j.total_cost:,.0f}\u0e3f")
            
            if not unpaid_rooms and not unpaid_houses and not unpaid_jobs:
                lines.append("\n\u2705 ไม่มีบิลค้างชำระครับ! เก็บเงินครบทุกรายการแล้ว \U0001f389")
            
            reply_text = "\n".join(lines)
            
            await get_line_bot_api().reply_message(
                ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=reply_text)]
                )
            )
            continue

        # Command D: รับ/จ่าย [amount] [description]
        match = re.match(r'^(รับ|จ่าย)\s+(\d+(?:\.\d+)?)\s+(.*)$', text, re.IGNORECASE)
        if match:
            action = match.group(1).lower()
            amount = float(match.group(2))
            description = match.group(3).strip()
            
            tx_type = models.TransactionType.INCOME if action == "รับ" else models.TransactionType.EXPENSE
            
            unit = match_business_unit(description, db)
            unit_id = unit.id if unit else None
            unit_name = unit.name if unit else "ทั่วไป"
            
            db_tx = models.Transaction(
                type=tx_type,
                amount=amount,
                description=description,
                unit_id=unit_id
            )
            db.add(db_tx)
            db.commit()
            db.refresh(db_tx)
            
            all_unit_tx = db.query(models.Transaction).filter(models.Transaction.unit_id == unit_id).all()
            total_inc = sum(t.amount for t in all_unit_tx if t.type == models.TransactionType.INCOME)
            total_exp = sum(t.amount for t in all_unit_tx if t.type == models.TransactionType.EXPENSE)
            unit_balance = total_inc - total_exp
            
            emoji = "✅ บันทึกรายรับสำเร็จ!" if tx_type == models.TransactionType.INCOME else "✅ บันทึกรายจ่ายสำเร็จ!"
            type_label = "รายรับ" if tx_type == models.TransactionType.INCOME else "รายจ่าย"
            type_icon = "💵" if tx_type == models.TransactionType.INCOME else "💸"
            
            reply_text = (
                f"{emoji}\n"
                f"-----------------------\n"
                f"{type_icon} ประเภท: {type_label}\n"
                f"💰 จำนวนเงิน: {amount:,.2f} บาท\n"
                f"📝 รายละเอียด: {description}\n"
                f"🏢 ธุรกิจ: {unit_name}\n"
                f"-----------------------\n"
                f"📊 ยอดคงเหลือธุรกิจนี้: {unit_balance:,.2f} บาท"
            )
            
            await get_line_bot_api().reply_message(
                ReplyMessageRequest(
                    reply_token=event.reply_token,
                    messages=[TextMessage(text=reply_text)]
                )
            )
            continue

    return {"status": "ok"}

# ==========================================
# Billing Reminder API
# ==========================================
@app.post("/notify/billing-reminder")
def send_billing_reminder(db: Session = Depends(get_db)):
    unpaid_rooms = db.query(models.DormRoom).filter(
        models.DormRoom.payment_status != "paid",
        models.DormRoom.tenant != "",
        models.DormRoom.tenant != None
    ).all()
    unpaid_houses = db.query(models.RentalHouse).filter(
        models.RentalHouse.payment_status != "paid",
        models.RentalHouse.tenant_name != "",
        models.RentalHouse.tenant_name != None
    ).all()
    unpaid_jobs = db.query(models.GarageJob).filter(
        models.GarageJob.payment_status != "paid"
    ).all()
    
    return {
        "status": "success",
        "unpaid_rooms": len(unpaid_rooms),
        "unpaid_houses": len(unpaid_houses),
        "unpaid_jobs": len(unpaid_jobs),
        "total_unpaid": len(unpaid_rooms) + len(unpaid_houses) + len(unpaid_jobs),
        "rooms": [{"number": r.number, "dorm_key": r.dorm_key, "tenant": r.tenant} for r in unpaid_rooms],
        "houses": [{"name": h.name, "tenant": h.tenant_name} for h in unpaid_houses],
        "jobs": [{"plate": j.license_plate, "customer": j.customer_name, "cost": j.total_cost} for j in unpaid_jobs]
    }

# ==========================================
# Dormitory (หอพัก) Endpoints
# ==========================================
@app.get("/rooms/", response_model=List[schemas.DormRoom])
def read_rooms(db: Session = Depends(get_db)):
    return db.query(models.DormRoom).all()

@app.patch("/rooms/{dorm_key}/{number}/", response_model=schemas.DormRoom)
def update_room(dorm_key: str, number: str, room_update: schemas.DormRoomUpdate, db: Session = Depends(get_db)):
    room = db.query(models.DormRoom).filter(
        models.DormRoom.dorm_key == dorm_key,
        models.DormRoom.number == number
    ).first()
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    
    old_status = room.payment_status
    
    update_data = room_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(room, key, value)
    
    db.commit()
    db.refresh(room)
    
    # Auto-log to Sovereign Vault Ledger when payment transitions to "paid"
    if room.payment_status == "paid" and old_status != "paid" and room.tenant:
        total_bill = room.rate + room.water_cost + room.electric_cost + room.cleaning_fee + room.other_fee + room.fine_cost
        if total_bill > 0:
            unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.DORMITORY).first()
            unit_id = unit.id if unit else None
            
            current_month = datetime.now().strftime("%m/%Y")
            description = f"ค่าเช่าห้อง {room.number} ({room.dorm_key}) รอบเดือน {current_month} - {room.tenant}"
            
            # Check duplicate to avoid multiple charges
            existing_tx = db.query(models.Transaction).filter(
                models.Transaction.description == description
            ).first()
            
            if not existing_tx:
                db_tx = models.Transaction(
                    type=models.TransactionType.INCOME,
                    amount=total_bill,
                    description=description,
                    unit_id=unit_id
                )
                db.add(db_tx)
                db.commit()
                
    return room

@app.post("/rooms/rollover/")
def rollover_rooms(db: Session = Depends(get_db)):
    try:
        rooms = db.query(models.DormRoom).all()
        for r in rooms:
            r.water_meter = r.water_meter or 0.0
            r.electricity_meter = r.electricity_meter or 0.0
            r.water_cost = 0.0
            r.electric_cost = 0.0
            r.cleaning_fee = 0.0
            r.other_fee = 0.0
            r.late_days = 0
            r.fine_cost = 0.0
            r.payment_status = "pending"
            r.payment_date = ""
            r.remark = ""
            r.move_out = ""
            r.vacant = ""
        db.commit()
        return {"status": "success", "message": "ขึ้นรอบบิลใหม่และซิงค์มิเตอร์เรียบร้อย!"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/rooms/reset/")
def reset_rooms(db: Session = Depends(get_db)):
    try:
        db.query(models.DormRoom).delete()
        db.commit()
        # Re-seed
        seed_rooms_and_houses()
        return {"status": "success", "message": "ล้างข้อมูลห้องพักทั้งหมดเรียบร้อยแล้ว!"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ==========================================
# Garage (อู่ซ่อมรถ) Endpoints
# ==========================================
@app.get("/garage/jobs/", response_model=List[schemas.GarageJob])
def read_garage_jobs(db: Session = Depends(get_db)):
    return db.query(models.GarageJob).order_by(models.GarageJob.created_at.desc()).all()

@app.post("/garage/jobs/", response_model=schemas.GarageJob)
def create_garage_job(job: schemas.GarageJobCreate, db: Session = Depends(get_db)):
    db_job = models.GarageJob(**job.dict())
    db.add(db_job)
    db.commit()
    db.refresh(db_job)
    return db_job

@app.patch("/garage/jobs/{job_id}/", response_model=schemas.GarageJob)
def update_garage_job(job_id: int, job_update: schemas.GarageJobUpdate, db: Session = Depends(get_db)):
    db_job = db.query(models.GarageJob).filter(models.GarageJob.id == job_id).first()
    if not db_job:
        raise HTTPException(status_code=404, detail="Garage job not found")
    
    old_payment_status = db_job.payment_status
    
    update_data = job_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_job, key, value)
    
    # Automate finished_at if status becomes finished or picked_up
    if db_job.status in ["finished", "picked_up"] and not db_job.finished_at:
        db_job.finished_at = datetime.utcnow()
    
    db.commit()
    db.refresh(db_job)
    
    # Auto-log to Sovereign Vault Ledger when payment transitions to "paid"
    if db_job.payment_status == "paid" and old_payment_status != "paid":
        if db_job.total_cost > 0:
            unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.GARAGE).first()
            unit_id = unit.id if unit else None
            
            description = f"ค่าซ่อมรถทะเบียน {db_job.license_plate} ({db_job.customer_name})"
            
            existing_tx = db.query(models.Transaction).filter(
                models.Transaction.description == description
            ).first()
            
            if not existing_tx:
                db_tx = models.Transaction(
                    type=models.TransactionType.INCOME,
                    amount=db_job.total_cost,
                    description=description,
                    unit_id=unit_id
                )
                db.add(db_tx)
                db.commit()
                
    return db_job

@app.delete("/garage/jobs/{job_id}/")
def delete_garage_job(job_id: int, db: Session = Depends(get_db)):
    db_job = db.query(models.GarageJob).filter(models.GarageJob.id == job_id).first()
    if not db_job:
        raise HTTPException(status_code=404, detail="Garage job not found")
    db.delete(db_job)
    db.commit()
    return {"status": "success", "message": f"Job {job_id} deleted successfully"}


# ==========================================
# Rental House (บ้านเช่า) Endpoints
# ==========================================
@app.get("/houses/", response_model=List[schemas.RentalHouse])
def read_rental_houses(db: Session = Depends(get_db)):
    return db.query(models.RentalHouse).all()

@app.patch("/houses/{house_id}/", response_model=schemas.RentalHouse)
def update_rental_house(house_id: str, house_update: schemas.RentalHouseUpdate, db: Session = Depends(get_db)):
    db_house = db.query(models.RentalHouse).filter(models.RentalHouse.id == house_id).first()
    if not db_house:
        raise HTTPException(status_code=404, detail="Rental house not found")
    
    old_payment_status = db_house.payment_status
    
    update_data = house_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_house, key, value)
    
    db.commit()
    db.refresh(db_house)
    
    # Auto-log to Sovereign Vault Ledger when payment transitions to "paid"
    if db_house.payment_status == "paid" and old_payment_status != "paid" and db_house.tenant_name:
        total_cost = db_house.monthly_rent + db_house.water_bill + db_house.electric_bill
        if total_cost > 0:
            # Query specific business unit based on name
            unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.name == db_house.name).first()
            if not unit:
                unit = db.query(models.BusinessUnit).filter(models.BusinessUnit.type == models.UnitType.HOUSE).first()
            unit_id = unit.id if unit else None
            
            current_month = datetime.now().strftime("%m/%Y")
            description = f"ค่าเช่าบ้าน {db_house.name} รอบเดือน {current_month} ({db_house.tenant_name})"
            
            existing_tx = db.query(models.Transaction).filter(
                models.Transaction.description == description
            ).first()
            
            if not existing_tx:
                db_tx = models.Transaction(
                    type=models.TransactionType.INCOME,
                    amount=total_cost,
                    description=description,
                    unit_id=unit_id
                )
                db.add(db_tx)
                db.commit()
                
    return db_house


